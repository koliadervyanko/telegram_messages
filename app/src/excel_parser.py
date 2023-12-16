import openpyxl

from .dto.excel_data_dto import ExcelDataDto


class ExcelParser:
    def __init__(self, file_name: str):
        self.__file_name = file_name
        self.__wb = openpyxl.load_workbook(self.__file_name)
        self.__ws = self.__wb["sheet"]

    def parse(self):
        links = self.__get_links()
        key_words = self.__get_key_words()
        return ExcelDataDto(links, key_words)

    def __get_links(self):
        links = []
        row = 2

        while True:
            cell_value = self.__ws.cell(row=row, column=1).value
            if cell_value is None:
                break
            else:
                links.append(cell_value)
                row += 1
        return links

    def __get_key_words(self):
        key_words = []
        row = 2
        while True:
            cell_value = self.__ws.cell(row=row, column=2).value
            if cell_value is None:
                break
            else:
                key_words.append(cell_value)
                row += 1
        return key_words
