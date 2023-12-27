import datetime
import sys

import openpyxl

from src.dto.excel_data_dto import ExcelDataDto


class ExcelParser:
    def __init__(self, file_name: str):
        self.__file_name = file_name
        self.__wb = openpyxl.load_workbook(self.__file_name)
        self.__ws = self.__wb["sheet"]

    def parse(self):
        links = self.__get_links()
        key_words = self.__get_key_words()
        date = self.__get_date()
        print(f"{self.__file_name} parsed successfully")
        return ExcelDataDto(links, key_words, date)

    def __get_date(self):
        data = self.__ws.cell(row=2, column=3).value
        parts = data.split("/")
        if data is None:
            print("Date is required")
            sys.exit(0)
        date = datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
        return date

    def __get_key_words(self):
        key_words = []
        row = 2
        while True:
            cell_value = self.__ws.cell(row=row, column=2).value
            if row == 2 and cell_value is None:
                print("Minimum number of keywords 1")
                sys.exit(0)
            if cell_value is None:
                break
            else:
                key_words.append(cell_value)
                row += 1
        return key_words

    def __get_links(self):
        links = []
        row = 2

        while True:
            cell_value = self.__ws.cell(row=row, column=1).value
            if row == 2 and cell_value is None:
                print("Minimum number of links 1")
                sys.exit(0)
            if cell_value is None:
                break
            else:
                links.append(cell_value)
                row += 1
        return links
